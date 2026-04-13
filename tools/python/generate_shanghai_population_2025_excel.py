from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 48)


def header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def make_title(ws, title):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)


wb = Workbook()

# Sheet 1: 2025 release status
ws1 = wb.active
ws1.title = "2025发布状态"
make_title(ws1, "2025年上海人口官方数据发布状态（截至2026-03-27）")
rows1 = [
    ["事项", "状态", "说明"],
    ["2025年上海市国民经济和社会发展统计公报", "待检索到公开正文", "2026年发布计划显示统计公报在3月中旬发布；在已抓取公开页面中未获取到该公报正文链接。"],
    ["2025年上海市国民经济运行情况", "已发布", "已发布于2026-01-21，但正文未披露人口总量与结构细项。"],
    ["2025年全国1%人口抽样调查（上海）", "进行中", "官方时间安排显示“公布和开发数据”为2026.04-2027.12。"],
    ["可用于当前分析的最新完整人口结构数据", "2024年", "来自2024年上海市国民经济和社会发展统计公报、2025上海统计年鉴（收录至2024年）。"],
]
for r in rows1:
    ws1.append(r)
for c in ws1[2]:
    header_style(c)
auto_width(ws1)

# Sheet 2: latest official population data
ws2 = wb.create_sheet("最新人口数据")
make_title(ws2, "上海最新官方人口数据（常住口径）")
rows2 = [
    ["年份", "常住人口(万人)", "户籍常住人口(万人)", "外来常住人口(万人)", "出生人数(万人)", "出生率(‰)", "死亡人数(万人)", "死亡率(‰)", "自然增长率(‰)", "出生性别比"],
    [2023, 2487.45, 1480.17, 1007.28, 9.8, 3.95, 15.8, 6.37, -2.42, 107.3],
    [2024, 2480.26, 1496.77, 983.49, 11.8, 4.75, 15.6, 6.28, -1.53, 107.2],
]
for r in rows2:
    ws2.append(r)
for c in ws2[2]:
    header_style(c)
auto_width(ws2)

# Sheet 3: structure data
ws3 = wb.create_sheet("人口结构(2024)")
make_title(ws3, "2024年人口结构（可得的最新官方结构数据）")
rows3 = [
    ["结构类型", "指标", "人数(万人)", "占比(%)", "口径说明"],
    ["常住人口构成", "户籍常住人口", 1496.77, round(1496.77 / 2480.26 * 100, 2), "2024统计公报"],
    ["常住人口构成", "外来常住人口", 983.49, round(983.49 / 2480.26 * 100, 2), "2024统计公报"],
    ["户籍人口年龄构成", "17岁及以下", 199.56, round(199.56 / 1535.09 * 100, 2), "2025上海统计年鉴 表2.6"],
    ["户籍人口年龄构成", "18-34岁", 210.91, round(210.91 / 1535.09 * 100, 2), "2025上海统计年鉴 表2.6"],
    ["户籍人口年龄构成", "35-59岁", 547.72, round(547.72 / 1535.09 * 100, 2), "2025上海统计年鉴 表2.6"],
    ["户籍人口年龄构成", "60岁及以上", 576.90, round(576.90 / 1535.09 * 100, 2), "2025上海统计年鉴 表2.6"],
    ["户籍老年人口内部结构", "60-64岁", 125.48, round(125.48 / 577.62 * 100, 2), "2025上海统计年鉴 表2.7"],
    ["户籍老年人口内部结构", "65-79岁", 366.14, round(366.14 / 577.62 * 100, 2), "2025上海统计年鉴 表2.7"],
    ["户籍老年人口内部结构", "80岁及以上", 86.00, round(86.00 / 577.62 * 100, 2), "2025上海统计年鉴 表2.7"],
]
for r in rows3:
    ws3.append(r)
for c in ws3[2]:
    header_style(c)
auto_width(ws3)

# Sheet 4: changes
ws4 = wb.create_sheet("变化分析")
make_title(ws4, "变化情况（2023→2024，基于最新可得官方数据）")
changes = [
    ["指标", "2023", "2024", "绝对变化", "相对变化(%)", "变化解读"],
]

pairs = [
    ("常住人口(万人)", 2487.45, 2480.26, "总量小幅回落"),
    ("户籍常住人口(万人)", 1480.17, 1496.77, "户籍常住人口增加"),
    ("外来常住人口(万人)", 1007.28, 983.49, "外来常住人口减少"),
    ("出生人数(万人)", 9.8, 11.8, "出生人数回升"),
    ("出生率(‰)", 3.95, 4.75, "出生率回升"),
    ("死亡人数(万人)", 15.8, 15.6, "死亡人数略降"),
    ("死亡率(‰)", 6.37, 6.28, "死亡率略降"),
    ("自然增长率(‰)", -2.42, -1.53, "自然负增长收窄"),
]

for name, v23, v24, note in pairs:
    diff = round(v24 - v23, 2)
    pct = round((diff / v23) * 100, 2) if v23 != 0 else None
    changes.append([name, v23, v24, diff, pct, note])

for r in changes:
    ws4.append(r)
for c in ws4[2]:
    header_style(c)
auto_width(ws4)

# Sheet 5: sources
ws5 = wb.create_sheet("数据来源")
make_title(ws5, "官方数据来源")
rows5 = [
    ["来源名称", "链接"],
    ["2025年上海市国民经济运行情况（上海市统计局）", "http://tjj.sh.gov.cn/tjxw/20260120/d3f4918a77484c77bed173db6aaef33c.html"],
    ["2024年上海市国民经济和社会发展统计公报（上海市统计局）", "https://tjj.sh.gov.cn/tjgb/20250324/a7fe18c6d5c24d66bfca89c5bb4cdcfb.html"],
    ["2023年上海市国民经济和社会发展统计公报（上海市统计局）", "https://tjj.sh.gov.cn/tjgb/20240321/f66c5b25ce604a1f9af755941d5f454a.html"],
    ["2025上海统计年鉴 表2.1（户数、人口、人口密度）", "https://tjj.sh.gov.cn/tjnj/nj25.htm?d1=2025tjnj/C0201.htm"],
    ["2025上海统计年鉴 表2.3（户籍人口出生率、死亡率、自然增长率）", "https://tjj.sh.gov.cn/tjnj/nj25.htm?d1=2025tjnj/C0203.htm"],
    ["2025上海统计年鉴 表2.6（各区户籍人口年龄构成）", "https://tjj.sh.gov.cn/tjnj/nj25.htm?d1=2025tjnj/C0206.htm"],
    ["2025上海统计年鉴 表2.7（各区户籍老年人口年龄构成）", "https://tjj.sh.gov.cn/tjnj/nj25.htm?d1=2025tjnj/C0207.htm"],
    ["上海市2025年全国1%人口抽样调查专题（进度）", "https://tjj.sh.gov.cn/2025rkcydc/"],
    ["2026年统计数据信息发布计划（统计公报发布时间）", "https://tjj.sh.gov.cn/fbjh/index.html"],
]
for r in rows5:
    ws5.append(r)
for c in ws5[2]:
    header_style(c)
auto_width(ws5)

output = "d:/VS project/上海市人口数据总结_2025版_官方口径.xlsx"
wb.save(output)
print(output)
